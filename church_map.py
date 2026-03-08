#!/usr/bin/env python3
"""
Church Attendance Probability Map & Chart
FBCJ (Jackson) vs Dalton second location

Geocodes ~318 unique addresses from the ServantKeeper directory,
calculates driving distances to both church locations, computes
attendance probability scores, and generates:
  - FBCJ_Church_Map.html  (interactive folium map with smooth contour boundary lines)
  - FBCJ_Church_Probability_Distance.html  (interactive Plotly chart by driving distance)
  - FBCJ_Church_Probability_Time.html  (interactive Plotly chart by driving time)
  - geocoded_addresses.json  (cached geocoding results)
"""

import json
import math
import os
import time
from pathlib import Path

import folium
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import plotly.graph_objects as go
import requests

# ── Configuration ──────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
INPUT_FILE = SCRIPT_DIR / "FBCJ_Directory #2.xlsx"
CACHE_FILE = SCRIPT_DIR / "geocoded_addresses.json"
MAP_OUTPUT = SCRIPT_DIR / "FBCJ_Church_Map.html"
MAP_OUTPUT_TIME = SCRIPT_DIR / "FBCJ_Church_Map_Time.html"
MAP_OUTPUT_DIST = SCRIPT_DIR / "FBCJ_Church_Map_Distance.html"
CHART_OUTPUT_DIST = SCRIPT_DIR / "FBCJ_Church_Probability_Distance.html"
CHART_OUTPUT_TIME = SCRIPT_DIR / "FBCJ_Church_Probability_Time.html"
CHART_OUTPUT_HIST = SCRIPT_DIR / "FBCJ_Church_Histogram.html"

# Church locations
JACKSON = {"name": "First Baptist Church of Jackson", "lat": 40.89482, "lon": -81.51603,
           "address": "8925 Strausser St NW, Massillon, OH 44646"}
DALTON = {"name": "Dalton Dariette (Proposed)", "lat": 40.79872, "lon": -81.69517,
          "address": "17 W Main St, Dalton, OH 44618"}

# States to skip (out of region)
SKIP_STATES = {"MX", "FL", "PA", "MI"}

# Geocoding settings
CENSUS_GEOCODER_URL = "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
NOMINATIM_HEADERS = {"User-Agent": "FBCJ-ChurchPlant-Analysis/1.0 (private church planning tool)"}

# Routing engine: "osrm" (free, no key) or "google" (requires valid API key)
ROUTING_ENGINE = "google"

# OSRM (OpenStreetMap routing) settings — free, no API key
OSRM_URL = "http://router.project-osrm.org/route/v1/driving"

# Google Maps Distance Matrix settings (if ROUTING_ENGINE == "google")
GOOGLE_MAPS_API_KEY = "AIzaSyCwEUbQGoWgJCha4QfuKMME1C-qypiiKU4"
GOOGLE_DISTANCE_MATRIX_URL = "https://maps.googleapis.com/maps/api/distancematrix/json"

# Boundary grid settings — centered on church midpoint, not on addresses
GRID_N = 40  # 40x40 = 1,600 points (~1.0 mi cells)
GRID_RADIUS_MI = 20  # symmetric radius around church midpoint
_CHURCH_MID_LAT = (JACKSON["lat"] + DALTON["lat"]) / 2
_CHURCH_MID_LON = (JACKSON["lon"] + DALTON["lon"]) / 2
# Convert miles to approximate degrees (1° lat ≈ 69 mi, 1° lon ≈ 53 mi at this latitude)
GRID_LAT_MIN = _CHURCH_MID_LAT - GRID_RADIUS_MI / 69.0
GRID_LAT_MAX = _CHURCH_MID_LAT + GRID_RADIUS_MI / 69.0
GRID_LON_MIN = _CHURCH_MID_LON - GRID_RADIUS_MI / 53.0
GRID_LON_MAX = _CHURCH_MID_LON + GRID_RADIUS_MI / 53.0

# Logistic model parameter — how sharply distance difference matters
# Higher k = sharper transition
K_LOGISTIC = 0.5  # per mile
K_LOGISTIC_TIME = 0.15  # per minute


# ── Step 1: Read and extract unique addresses ──────────────────────

# Street abbreviation map for address normalization
_STREET_ABBREVS = {
    "dr": "drive", "dr.": "drive",
    "st": "street", "st.": "street",
    "ave": "avenue", "ave.": "avenue",
    "rd": "road", "rd.": "road",
    "ln": "lane", "ln.": "lane",
    "ct": "court", "ct.": "court",
    "cir": "circle", "cir.": "circle",
    "blvd": "boulevard", "blvd.": "boulevard",
    "pl": "place", "pl.": "place",
    "pkwy": "parkway", "pkwy.": "parkway",
    "hwy": "highway", "hwy.": "highway",
    "trl": "trail", "trl.": "trail",
}


def _normalize_street(street):
    """Normalize a street string for dedup keying.

    Strips directional suffixes (NW/NE/SW/SE), expands common abbreviations,
    removes punctuation, and lowercases. Used only for matching — the original
    street string is preserved for geocoding and display.
    """
    import re
    s = street.strip()
    # Remove trailing directional suffixes (with optional comma, must be standalone word)
    s = re.sub(r',?\s+(?:NW|NE|SW|SE)\s*$', '', s, flags=re.IGNORECASE).strip()
    # Remove trailing periods and commas
    s = s.rstrip('.,').strip()
    # Lowercase for comparison
    s = s.lower()
    # Expand abbreviations (match whole words only)
    words = s.split()
    words = [_STREET_ABBREVS.get(w.rstrip('.,'), w) for w in words]
    s = ' '.join(words)
    # Remove all remaining periods and extra whitespace
    s = s.replace('.', '').replace(',', '')
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _classify_member(grade):
    """Classify a person as adult, child, or graduated based on School Grade field."""
    if not grade:
        return "adult"
    g = str(grade).strip().lower()
    if g in ("", "none"):
        return "adult"
    if g == "graduated":
        return "graduated"
    # Everything else (K, 1st-12th, Kindergarten, etc.) = child
    return "child"


def read_addresses():
    """Read FBCJ directory and return unique addresses with family names and member details."""
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb["FBCJ Directory"]

    addresses = {}  # key: norm_key -> {"families": [], "members": [], "street": original, ...}
    skipped = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        family_name = row[1]
        ind_name = row[2]
        street = row[6]
        city = row[8]
        state = row[9]
        zipcode = row[10]
        grade = row[18] if len(row) > 18 else None

        if not street or not city:
            continue

        if state in SKIP_STATES:
            skipped.append({"family": family_name, "address": f"{street}, {city}, {state} {zipcode}",
                            "reason": f"Out of region ({state})"})
            continue

        street_raw = street.strip()
        city_clean = city.strip()
        state_clean = (state or "OH").strip()
        zip_clean = str(zipcode or "").strip()

        norm_key = (_normalize_street(street_raw), city_clean.lower(), state_clean, zip_clean)

        if norm_key not in addresses:
            addresses[norm_key] = {"families": [], "members": [], "street": street_raw,
                                    "city": city_clean, "state": state_clean, "zip": zip_clean}
        if family_name and family_name not in addresses[norm_key]["families"]:
            addresses[norm_key]["families"].append(family_name)
        # Keep the longest/most-complete street string for geocoding
        if len(street_raw) > len(addresses[norm_key]["street"]):
            addresses[norm_key]["street"] = street_raw

        # Track individual members with their classification
        member_name = str(ind_name).strip() if ind_name else str(family_name).strip() if family_name else "Unknown"
        member_type = _classify_member(grade)
        addresses[norm_key]["members"].append({"name": member_name, "type": member_type,
                                                "grade": str(grade).strip() if grade else None})

    # Convert to the format the rest of the pipeline expects: {(street, city, state, zip): [families]}
    # Also build a parallel members dict keyed the same way
    result = {}
    members_by_address = {}
    merged_count = 0
    for norm_key, data in addresses.items():
        out_key = (data["street"], data["city"], data["state"], data["zip"])
        if out_key in result:
            for fam in data["families"]:
                if fam not in result[out_key]:
                    result[out_key].append(fam)
            members_by_address[out_key].extend(data["members"])
        else:
            result[out_key] = data["families"]
            members_by_address[out_key] = data["members"]

    raw_count = len(addresses)
    if raw_count != len(result):
        merged_count = raw_count - len(result)

    # Count totals for summary
    all_members = [m for members in members_by_address.values() for m in members]
    adults = sum(1 for m in all_members if m["type"] == "adult")
    children = sum(1 for m in all_members if m["type"] == "child")
    graduated = sum(1 for m in all_members if m["type"] == "graduated")

    print(f"Found {len(result)} unique OH addresses (merged {merged_count} near-duplicates), "
          f"skipped {len(skipped)} out-of-region")
    print(f"  Members: {adults} adults, {graduated} graduated, {children} children (K-12)")
    if skipped:
        seen = set()
        for s in skipped:
            if s["family"] not in seen:
                print(f"  Skipped: {s['family']} — {s['address']}")
                seen.add(s["family"])
    return result, members_by_address


# ── Step 2: Geocode addresses ──────────────────────────────────────

def _geocode_census(address_line):
    """Geocode using US Census Bureau geocoder. Returns (lat, lon) or None."""
    try:
        resp = requests.get(CENSUS_GEOCODER_URL, params={
            "address": address_line,
            "benchmark": "Public_AR_Current",
            "format": "json"
        }, timeout=15)
        matches = resp.json().get("result", {}).get("addressMatches", [])
        if matches:
            coords = matches[0]["coordinates"]
            return float(coords["y"]), float(coords["x"])
    except Exception:
        pass
    return None


def _geocode_nominatim_zip(zipcode):
    """Fallback: get zip code centroid from Nominatim."""
    try:
        resp = requests.get(NOMINATIM_URL, params={
            "postalcode": str(zipcode), "country": "US",
            "format": "json", "limit": 1
        }, headers=NOMINATIM_HEADERS, timeout=10)
        data = resp.json()
        if data:
            return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception:
        pass
    return None


def geocode_addresses(addresses, members_by_address=None):
    """Geocode addresses using US Census Bureau (primary) + Nominatim zip fallback."""
    import re
    import sys

    if members_by_address is None:
        members_by_address = {}

    cache = {}
    if CACHE_FILE.exists():
        with open(CACHE_FILE) as f:
            cache = json.load(f)
        print(f"Loaded {len(cache)} cached geocoding results", flush=True)

    results = []
    to_geocode = []

    for (street, city, state, zipcode), families in addresses.items():
        addr_key = (street, city, state, zipcode)
        cache_key = f"{street}|{city}|{state}|{zipcode}"
        members = members_by_address.get(addr_key, [])
        if cache_key in cache:
            entry = cache[cache_key].copy()
            entry["families"] = families
            entry["members"] = members
            results.append(entry)
        else:
            to_geocode.append((street, city, state, zipcode, families, cache_key, members))

    if to_geocode:
        print(f"Geocoding {len(to_geocode)} new addresses via Census Bureau...", flush=True)
        failed = []
        exact_count = 0
        approx_count = 0

        for i, (street, city, state, zipcode, families, cache_key, members) in enumerate(to_geocode):
            address_line = f"{street}, {city}, {state or 'OH'} {zipcode}"
            approximate = False

            # Primary: US Census Bureau geocoder
            result = _geocode_census(address_line)

            # Fallback: try without directional suffix
            if result is None:
                cleaned = re.sub(r',?\s*(NW|NE|SW|SE)\s*$', '', street, flags=re.IGNORECASE).strip()
                if cleaned != street:
                    result = _geocode_census(f"{cleaned}, {city}, {state or 'OH'} {zipcode}")
                    time.sleep(0.3)

            # Fallback: zip code centroid (approximate)
            if result is None and zipcode:
                result = _geocode_nominatim_zip(zipcode)
                if result:
                    approximate = True
                time.sleep(1)  # Nominatim rate limit

            if result:
                lat, lon = result
                entry = {
                    "street": street, "city": city, "state": state, "zip": zipcode,
                    "lat": lat, "lon": lon,
                    "families": families,
                    "members": members,
                    "approximate": approximate
                }
                cache[cache_key] = {k: v for k, v in entry.items() if k not in ("families", "members")}
                results.append(entry)
                if approximate:
                    approx_count += 1
                else:
                    exact_count += 1
            else:
                failed.append(address_line)

            # Census doesn't require 1s delay, but be polite
            time.sleep(0.2)

            if (i + 1) % 20 == 0:
                print(f"  Geocoded {i + 1}/{len(to_geocode)} ({exact_count} exact, {approx_count} approx)...", flush=True)

        print(f"\n  Results: {exact_count} exact, {approx_count} approximate (zip centroid)", flush=True)

        if failed:
            print(f"  Failed completely: {len(failed)}", flush=True)
            for f in failed[:5]:
                print(f"    - {f}", flush=True)

        with open(CACHE_FILE, "w") as f:
            json.dump(cache, f, indent=2)
        print(f"  Saved {len(cache)} total geocoding results to cache", flush=True)
    else:
        print("All addresses already cached!", flush=True)

    return results


# ── Step 3: Calculate driving distances ────────────────────────────

def haversine_miles(lat1, lon1, lat2, lon2):
    """Calculate straight-line distance in miles between two points."""
    R = 3959  # Earth radius in miles
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))


def _get_osrm_distance(lat, lon, dest_lat, dest_lon):
    """Get driving distance (miles) and time (minutes) via OSRM (free, no key)."""
    try:
        url = f"{OSRM_URL}/{lon},{lat};{dest_lon},{dest_lat}"
        resp = requests.get(url, params={"overview": "false"}, timeout=10)
        data = resp.json()
        if data.get("code") == "Ok" and data.get("routes"):
            route = data["routes"][0]
            miles = route["distance"] / 1609.34  # meters to miles
            minutes = route["duration"] / 60  # seconds to minutes
            return miles, minutes
    except Exception:
        pass
    return None, None


def _next_sunday_9am_epoch():
    """Return Unix timestamp for next Sunday at 9:00 AM Eastern."""
    from datetime import datetime, timedelta
    import calendar
    now = datetime.utcnow()
    # Eastern time is UTC-5 (EST) or UTC-4 (EDT); use UTC-4 for March
    eastern_9am_utc = 13  # 9 AM EDT = 13:00 UTC
    days_until_sunday = (calendar.SUNDAY - now.weekday()) % 7
    if days_until_sunday == 0:
        days_until_sunday = 7  # always pick next Sunday, not today
    next_sun = now.replace(hour=eastern_9am_utc, minute=0, second=0, microsecond=0) + timedelta(days=days_until_sunday)
    return int(next_sun.timestamp())


def _get_google_distance(lat, lon, dest_lat, dest_lon):
    """Get driving distance (miles) and time (minutes) via Google Distance Matrix API."""
    try:
        resp = requests.get(GOOGLE_DISTANCE_MATRIX_URL, params={
            "origins": f"{lat},{lon}",
            "destinations": f"{dest_lat},{dest_lon}",
            "departure_time": _next_sunday_9am_epoch(),
            "key": GOOGLE_MAPS_API_KEY,
        }, timeout=10)
        data = resp.json()
        if data.get("status") == "OK":
            element = data["rows"][0]["elements"][0]
            if element.get("status") == "OK":
                miles = element["distance"]["value"] / 1609.34  # meters to miles
                minutes = element["duration"]["value"] / 60  # seconds to minutes
                return miles, minutes
    except Exception:
        pass
    return None, None


def get_google_distance(lat, lon, dest_lat, dest_lon):
    """Get driving distance/time using configured routing engine."""
    if ROUTING_ENGINE == "osrm":
        return _get_osrm_distance(lat, lon, dest_lat, dest_lon)
    return _get_google_distance(lat, lon, dest_lat, dest_lon)


def calculate_distances(geocoded):
    """Calculate driving distances to both churches for each address."""
    print(f"\nCalculating driving distances for {len(geocoded)} addresses...")

    # Cache invalidation: detect routing engine switch (OSRM → Google)
    cache = {}
    if CACHE_FILE.exists():
        with open(CACHE_FILE) as f:
            cache = json.load(f)

    if cache.get("_routing_engine") != ROUTING_ENGINE:
        print(f"  Routing engine changed to {ROUTING_ENGINE} — invalidating cached distances...")
        dist_fields = {"dist_jackson_mi", "time_jackson_min", "dist_dalton_mi", "time_dalton_min",
                       "dist_diff", "time_diff", "p_jackson", "p_dalton", "closer_church"}
        for key, val in cache.items():
            if key.startswith("_"):
                continue
            if isinstance(val, dict):
                for field in dist_fields:
                    val.pop(field, None)
        cache.pop("_boundary_grid", None)
        cache["_routing_engine"] = ROUTING_ENGINE
        with open(CACHE_FILE, "w") as f:
            json.dump(cache, f, indent=2)
        # Strip distance fields from in-memory geocoded entries too
        for entry in geocoded:
            for field in dist_fields:
                entry.pop(field, None)

    # Check if distances are already cached in geocoded data
    need_calc = [g for g in geocoded if "dist_jackson_mi" not in g]
    if not need_calc:
        print("All distances already calculated!")
        return geocoded

    print(f"  Need to calculate {len(need_calc)} addresses ({len(need_calc) * 2} API calls)...")
    api_failures = 0

    for i, entry in enumerate(need_calc):
        lat, lon = entry["lat"], entry["lon"]

        # Distance to Jackson
        jm, jt = get_google_distance(lat, lon, JACKSON["lat"], JACKSON["lon"])
        if jm is not None:
            entry["dist_jackson_mi"] = round(jm, 1)
            entry["time_jackson_min"] = round(jt, 1)
        else:
            # Fallback to straight-line * 1.3 (road factor)
            sl = haversine_miles(lat, lon, JACKSON["lat"], JACKSON["lon"])
            entry["dist_jackson_mi"] = round(sl * 1.3, 1)
            entry["time_jackson_min"] = round(sl * 1.3 * 2, 1)  # ~30mph avg
            api_failures += 1
        time.sleep(0.3)

        # Distance to Dalton
        dm, dt = get_google_distance(lat, lon, DALTON["lat"], DALTON["lon"])
        if dm is not None:
            entry["dist_dalton_mi"] = round(dm, 1)
            entry["time_dalton_min"] = round(dt, 1)
        else:
            sl = haversine_miles(lat, lon, DALTON["lat"], DALTON["lon"])
            entry["dist_dalton_mi"] = round(sl * 1.3, 1)
            entry["time_dalton_min"] = round(sl * 1.3 * 2, 1)
            api_failures += 1
        time.sleep(0.3)

        if (i + 1) % 20 == 0:
            print(f"  Distances: {i + 1}/{len(need_calc)}...")

    if api_failures:
        print(f"  {api_failures} Google API failures, used straight-line fallback")

    # Update cache with distances
    cache = {}
    if CACHE_FILE.exists():
        with open(CACHE_FILE) as f:
            cache = json.load(f)

    for entry in geocoded:
        cache_key = f"{entry['street']}|{entry['city']}|{entry['state']}|{entry['zip']}"
        cache[cache_key] = {k: v for k, v in entry.items() if k not in ("families", "members")}

    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)

    return geocoded


def retry_fallback_distances(geocoded):
    """Retry addresses that fell back to straight-line estimates."""
    fallbacks = []
    for entry in geocoded:
        # Detect straight-line fallback: time ≈ distance * 2
        if abs(entry.get("time_jackson_min", 0) - entry.get("dist_jackson_mi", 0) * 2) < 0.2:
            fallbacks.append(entry)

    if not fallbacks:
        print("\nNo straight-line fallback entries to retry")
        return geocoded

    print(f"\nRetrying {len(fallbacks)} addresses that used straight-line fallback...")
    fixed = 0
    for entry in fallbacks:
        lat, lon = entry["lat"], entry["lon"]

        jm, jt = _get_google_distance(lat, lon, JACKSON["lat"], JACKSON["lon"])
        time.sleep(0.3)
        dm, dt = _get_google_distance(lat, lon, DALTON["lat"], DALTON["lon"])
        time.sleep(0.3)

        if jm is not None and dm is not None:
            entry["dist_jackson_mi"] = round(jm, 1)
            entry["time_jackson_min"] = round(jt, 1)
            entry["dist_dalton_mi"] = round(dm, 1)
            entry["time_dalton_min"] = round(dt, 1)
            fixed += 1

    print(f"  Fixed {fixed}/{len(fallbacks)} entries with real Google data")

    # Update cache
    if fixed:
        cache = {}
        if CACHE_FILE.exists():
            with open(CACHE_FILE) as f:
                cache = json.load(f)
        for entry in geocoded:
            cache_key = f"{entry['street']}|{entry['city']}|{entry['state']}|{entry['zip']}"
            cache[cache_key] = {k: v for k, v in entry.items() if k not in ("families", "members")}
        with open(CACHE_FILE, "w") as f:
            json.dump(cache, f, indent=2)

    return geocoded


# ── Step 4: Calculate probability scores ───────────────────────────

def filter_geographic_outliers(geocoded, max_miles=35):
    """Remove addresses more than max_miles from the midpoint of the two churches."""
    center_lat = (JACKSON["lat"] + DALTON["lat"]) / 2
    center_lon = (JACKSON["lon"] + DALTON["lon"]) / 2

    filtered = []
    excluded = []
    for entry in geocoded:
        dist = haversine_miles(entry["lat"], entry["lon"], center_lat, center_lon)
        if dist <= max_miles:
            filtered.append(entry)
        else:
            excluded.append(entry)

    if excluded:
        print(f"\nGeographic filter: excluded {len(excluded)} addresses > {max_miles} mi from church midpoint:")
        for e in excluded:
            families = ", ".join(e.get("families", ["Unknown"]))
            dist = haversine_miles(e["lat"], e["lon"], center_lat, center_lon)
            print(f"  - {families} ({e['city']}) — {dist:.1f} mi from midpoint")
    else:
        print(f"\nGeographic filter: all addresses within {max_miles} mi of midpoint")

    return filtered


def calculate_probabilities(geocoded):
    """Calculate P(Jackson) for each address using logistic model on driving time."""
    for entry in geocoded:
        tj = entry["time_jackson_min"]
        td = entry["time_dalton_min"]

        # Time-based: P(Jackson) = 1 / (1 + exp(k * (time_jackson - time_dalton)))
        time_diff = tj - td  # positive means Jackson takes longer
        p_jackson = 1.0 / (1.0 + math.exp(K_LOGISTIC_TIME * time_diff))

        entry["p_jackson"] = round(p_jackson, 3)
        entry["p_dalton"] = round(1 - p_jackson, 3)
        entry["dist_diff"] = round(entry["dist_jackson_mi"] - entry["dist_dalton_mi"], 1)
        entry["time_diff"] = round(time_diff, 1)
        entry["closer_church"] = "Jackson" if tj <= td else "Dalton"

    return geocoded


# ── Step 4b: Compute boundary grid ───────────────────────────────

def _compute_boundary_grid(lat_min=GRID_LAT_MIN, lat_max=GRID_LAT_MAX,
                           lon_min=GRID_LON_MIN, lon_max=GRID_LON_MAX, grid_n=GRID_N):
    """Compute driving distance/time to both churches on a grid via Google Maps.

    Results are cached in geocoded_addresses.json under '_boundary_grid' key.
    """
    cache = {}
    if CACHE_FILE.exists():
        with open(CACHE_FILE) as f:
            cache = json.load(f)

    cache_key = "_boundary_grid"
    expected = grid_n * grid_n
    if cache_key in cache:
        grid = cache[cache_key]
        # Invalidate if grid size doesn't match (e.g. old 20x20 vs new 25x25)
        if grid and len(grid) == expected:
            # Check if grid has actual data (not all nulls from API failures)
            has_data = any(pt.get("dist_jackson") is not None for pt in grid)
            if has_data:
                print(f"    Loaded {len(grid)} cached boundary grid points")
                return grid
            else:
                print(f"    Cached grid has all null values — recomputing with fallback...")
        elif grid:
            print(f"    Cached grid has {len(grid)} points, need {expected} — recomputing...")

    print(f"    Computing {grid_n}x{grid_n} boundary grid ({grid_n**2} points, ~{grid_n**2 * 2 * 0.3 / 60:.1f} min)...")
    grid_data = []
    total = grid_n * grid_n
    for i in range(grid_n):
        for j in range(grid_n):
            glat = lat_min + (lat_max - lat_min) * i / (grid_n - 1)
            glon = lon_min + (lon_max - lon_min) * j / (grid_n - 1)

            dj_mi, tj_min = get_google_distance(glat, glon, JACKSON["lat"], JACKSON["lon"])
            time.sleep(0.3)
            dd_mi, td_min = get_google_distance(glat, glon, DALTON["lat"], DALTON["lon"])
            time.sleep(0.3)

            # Fallback to straight-line estimates if API fails
            if dj_mi is None:
                sl = haversine_miles(glat, glon, JACKSON["lat"], JACKSON["lon"])
                dj_mi = sl * 1.3
                tj_min = sl * 1.3 * 2  # ~30mph avg
            if dd_mi is None:
                sl = haversine_miles(glat, glon, DALTON["lat"], DALTON["lon"])
                dd_mi = sl * 1.3
                td_min = sl * 1.3 * 2

            grid_data.append({
                "lat": round(glat, 5),
                "lon": round(glon, 5),
                "dist_jackson": round(dj_mi, 2),
                "dist_dalton": round(dd_mi, 2),
                "time_jackson": round(tj_min, 2),
                "time_dalton": round(td_min, 2),
            })

            done = i * grid_n + j + 1
            if done % 40 == 0 or done == total:
                print(f"    Grid: {done}/{total} points...", flush=True)

    cache[cache_key] = grid_data
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)
    print(f"    Cached {len(grid_data)} boundary grid points")
    return grid_data


# ── Step 5: Generate interactive map ───────────────────────────────

def probability_color(p_jackson):
    """Return a hex color: red (Jackson) → purple (toss-up) → blue (Dalton)."""
    # p_jackson: 1.0 = pure red, 0.0 = pure blue, 0.5 = purple
    r = int(200 * p_jackson + 50)
    b = int(200 * (1 - p_jackson) + 50)
    g = int(50 * (1 - abs(p_jackson - 0.5) * 2))  # peak green at 0.5 for purple
    return f"#{min(r,255):02x}{min(g,255):02x}{min(b,255):02x}"


def _add_church_markers(m):
    """Add church location markers to a folium map."""
    for church, color in [(JACKSON, "red"), (DALTON, "blue")]:
        folium.Marker(
            location=[church["lat"], church["lon"]],
            popup=f"<b>{church['name']}</b><br>{church['address']}",
            tooltip=church["name"],
            icon=folium.Icon(color=color, icon="home", prefix="fa")
        ).add_to(m)


def _add_address_markers(m, geocoded):
    """Add address markers with probability popups to a folium map."""
    jackson_group = folium.FeatureGroup(name="Lean Jackson (>60%)")
    dalton_group = folium.FeatureGroup(name="Lean Dalton (>60%)")
    tossup_group = folium.FeatureGroup(name="Toss-up (40-60%)")
    approx_group = folium.FeatureGroup(name="Low Confidence (approx/fallback)")

    for entry in geocoded:
        p = entry["p_jackson"]
        color = probability_color(p)
        families_str = ", ".join(entry.get("families", ["Unknown"]))

        time_diff = entry.get("time_diff", 0)
        time_diff_str = f"{abs(time_diff):.1f} min {'closer to Dalton' if time_diff > 0 else 'closer to Jackson'}"

        # Flag data quality issues
        is_approx = entry.get("approximate", False)
        is_fallback = abs(entry.get("time_jackson_min", 0) - entry.get("dist_jackson_mi", 0) * 2) < 0.2
        quality_notes = []
        if is_approx:
            quality_notes.append("ZIP centroid (approx location)")
        if is_fallback:
            quality_notes.append("Straight-line estimate (API failed)")
        quality_html = ""
        if quality_notes:
            quality_html = "<tr><td colspan='2' style='color:#cc6600;font-style:italic;padding-top:4px;'>" + "<br>".join(quality_notes) + "</td></tr>"

        popup_html = f"""
        <div style='width:280px'>
        <b>{families_str}</b><br>
        {entry['street']}, {entry['city']}<br><br>
        <table style='font-size:12px'>
        <tr><td>To Jackson:</td><td><b>{entry['dist_jackson_mi']} mi</b> ({entry['time_jackson_min']} min)</td></tr>
        <tr><td>To Dalton:</td><td><b>{entry['dist_dalton_mi']} mi</b> ({entry['time_dalton_min']} min)</td></tr>
        <tr><td>Time diff:</td><td><b>{time_diff_str}</b></td></tr>
        <tr><td>P(Jackson):</td><td><b>{entry['p_jackson']:.0%}</b></td></tr>
        <tr><td>P(Dalton):</td><td><b>{entry['p_dalton']:.0%}</b></td></tr>
        <tr><td>Closer:</td><td><b>{entry['closer_church']}</b></td></tr>
        {quality_html}
        </table>
        </div>
        """

        low_confidence = is_approx or is_fallback
        radius = 10 if not low_confidence else 8
        dash = None if not low_confidence else "4 3"

        marker = folium.CircleMarker(
            location=[entry["lat"], entry["lon"]],
            radius=radius,
            color=color if not low_confidence else "#999",
            fill=True,
            fill_color=color,
            fill_opacity=0.7 if not low_confidence else 0.4,
            popup=folium.Popup(popup_html, max_width=300),
            tooltip=f"{families_str} — P(Jackson)={p:.0%}" + (" [low confidence]" if low_confidence else ""),
            dash_array=dash,
        )

        if low_confidence:
            marker.add_to(approx_group)
        elif p > 0.6:
            marker.add_to(jackson_group)
        elif p < 0.4:
            marker.add_to(dalton_group)
        else:
            marker.add_to(tossup_group)

    jackson_group.add_to(m)
    dalton_group.add_to(m)
    tossup_group.add_to(m)
    approx_group.add_to(m)


def generate_map(geocoded):
    """Generate interactive folium map."""
    print("\nGenerating interactive map...")

    # Center map between the two churches
    center_lat = (JACKSON["lat"] + DALTON["lat"]) / 2
    center_lon = (JACKSON["lon"] + DALTON["lon"]) / 2

    m = folium.Map(location=[center_lat, center_lon], zoom_start=11,
                   tiles="OpenStreetMap")

    _add_church_markers(m)
    _add_address_markers(m, geocoded)

    # ── Dual decision boundary lines (driving distance + time) ──
    print("  Computing Google Maps decision boundaries...")

    grid_data = _compute_boundary_grid()

    # Build 2D grids for contour extraction
    grid_n = int(math.sqrt(len(grid_data)))
    lats_1d = sorted(set(pt["lat"] for pt in grid_data))
    lons_1d = sorted(set(pt["lon"] for pt in grid_data))

    # Create lookup for grid values
    grid_lookup = {(pt["lat"], pt["lon"]): pt for pt in grid_data}

    def _extract_contour(value_func, label, color):
        """Extract smooth contour at level=0 using matplotlib and return folium PolyLine."""
        Z = np.full((len(lats_1d), len(lons_1d)), np.nan)
        for i, lat in enumerate(lats_1d):
            for j, lon in enumerate(lons_1d):
                pt = grid_lookup.get((lat, lon))
                if pt:
                    val = value_func(pt)
                    if val is not None:
                        Z[i, j] = val

        # Fill NaNs for contour to work
        mask = np.isnan(Z)
        if mask.all():
            return None
        if mask.any():
            Z_filled = Z.copy()
            mean_val = np.nanmean(Z)
            Z_filled[mask] = mean_val
            Z = Z_filled

        # Interpolate to finer grid for smooth contours
        from scipy.ndimage import zoom, gaussian_filter
        Z = zoom(Z, 6, order=1)  # linear interp (no ringing artifacts)
        Z = gaussian_filter(Z, sigma=3)  # smooth out grid staircase
        lats_fine = np.linspace(lats_1d[0], lats_1d[-1], Z.shape[0])
        lons_fine = np.linspace(lons_1d[0], lons_1d[-1], Z.shape[1])

        # Extract contour at level 0
        fig_tmp, ax_tmp = plt.subplots()
        cs = ax_tmp.contour(lons_fine, lats_fine, Z, levels=[0])
        plt.close(fig_tmp)

        # Get contour paths via allsegs
        all_segments = []
        for segs in cs.allsegs:
            for seg in segs:
                segment = [[v[1], v[0]] for v in seg]  # [lat, lon] for folium
                if len(segment) >= 2:
                    all_segments.append(segment)

        if not all_segments:
            return None

        # Use the longest segment as the boundary
        boundary = max(all_segments, key=len)

        folium.PolyLine(
            boundary,
            color=color,
            weight=3,
            dash_array="10 6",
            opacity=0.8,
            tooltip=label
        ).add_to(m)

        mid = boundary[len(boundary) // 2]
        folium.Marker(
            location=mid,
            icon=folium.DivIcon(html=f'<div style="font-size:11px;font-weight:bold;color:{color};'
                                     f'background:rgba(255,255,255,0.85);padding:2px 6px;'
                                     f'border-radius:3px;white-space:nowrap;">'
                                     f'{label}</div>')
        ).add_to(m)
        return boundary

    # Distance boundary: contour where dist_jackson - dist_dalton = 0
    dist_line = _extract_contour(
        lambda pt: (pt["dist_jackson"] - pt["dist_dalton"]) if pt["dist_jackson"] and pt["dist_dalton"] else None,
        "Equal Driving Distance", "#cc3333"
    )
    if dist_line:
        print(f"    Distance boundary: {len(dist_line)} points (smooth contour)")
    else:
        print("    Warning: could not extract distance boundary contour")

    # Time boundary: contour where time_jackson - time_dalton = 0
    time_line = _extract_contour(
        lambda pt: (pt["time_jackson"] - pt["time_dalton"]) if pt["time_jackson"] and pt["time_dalton"] else None,
        "Equal Driving Time", "#3333cc"
    )
    if time_line:
        print(f"    Time boundary: {len(time_line)} points (smooth contour)")
    else:
        print("    Warning: could not extract time boundary contour")

    # Add legend
    # Count quality stats for legend
    hq_count = sum(1 for e in geocoded if e.get("is_high_quality", True))
    lq_count = len(geocoded) - hq_count

    legend_html = f"""
    <div style="position:fixed;bottom:30px;left:30px;z-index:1000;
                background:white;padding:12px 16px;border-radius:8px;
                box-shadow:0 2px 6px rgba(0,0,0,0.3);font-size:13px;max-width:300px;">
    <b>Attendance Probability</b><br>
    <span style="color:#fa3232;">&#9679;</span> Lean Jackson (&gt;60%)<br>
    <span style="color:#9932fa;">&#9679;</span> Toss-up (40–60%)<br>
    <span style="color:#3232fa;">&#9679;</span> Lean Dalton (&gt;60%)<br>
    <span style="color:#999;">&#9675;</span> Low Confidence ({lq_count} addresses)<br>
    <span style="color:#cc3333;">- - -</span> Equal Driving Distance<br>
    <span style="color:#3333cc;">- - -</span> Equal Driving Time<br>
    <hr style="margin:6px 0;border-color:#ddd;">
    <span style="font-size:11px;color:#666;">
    Model: logistic, K={K_LOGISTIC_TIME}/min<br>
    Bias: 0 (no incumbency factor)<br>
    Times: Google Maps, Sun 9AM EDT<br>
    <i>*EST months may vary ±1–2 min</i><br>
    {lq_count} addresses use approx. location<br>
    or straight-line distance estimates<br>
    See Sensitivity Analysis for robustness
    </span>
    </div>
    """
    m.get_root().html.add_child(folium.Element(legend_html))

    # Add layer control
    folium.LayerControl().add_to(m)

    m.save(str(MAP_OUTPUT))
    print(f"  Map saved to {MAP_OUTPUT}")

    # Print summary stats
    jackson_count = sum(1 for e in geocoded if e["p_jackson"] > 0.6)
    dalton_count = sum(1 for e in geocoded if e["p_jackson"] < 0.4)
    tossup_count = len(geocoded) - jackson_count - dalton_count
    print(f"\n  Summary: {jackson_count} lean Jackson, {dalton_count} lean Dalton, {tossup_count} toss-up")


# ── Step 5b: Generate graduated contour maps ──────────────────────

def generate_contour_map(geocoded, mode):
    """Generate a graduated contour map for time or distance differences.

    mode: "time" or "distance"
    """
    from matplotlib.colors import Normalize

    if mode == "time":
        output_path = MAP_OUTPUT_TIME
        unit_label = "min"
        title = "Driving Time Advantage"
        value_func = lambda pt: (pt["time_jackson"] - pt["time_dalton"]) if pt["time_jackson"] and pt["time_dalton"] else None
    else:
        output_path = MAP_OUTPUT_DIST
        unit_label = "mi"
        title = "Driving Distance Advantage"
        value_func = lambda pt: (pt["dist_jackson"] - pt["dist_dalton"]) if pt["dist_jackson"] and pt["dist_dalton"] else None

    print(f"\nGenerating {mode} contour map...")

    center_lat = (JACKSON["lat"] + DALTON["lat"]) / 2
    center_lon = (JACKSON["lon"] + DALTON["lon"]) / 2

    m = folium.Map(location=[center_lat, center_lon], zoom_start=11, tiles="OpenStreetMap")

    # Compute boundary grid (cached, centered on church midpoint)
    grid_data = _compute_boundary_grid()

    # Build 2D arrays
    lats_1d = sorted(set(pt["lat"] for pt in grid_data))
    lons_1d = sorted(set(pt["lon"] for pt in grid_data))
    grid_lookup = {(pt["lat"], pt["lon"]): pt for pt in grid_data}

    Z = np.full((len(lats_1d), len(lons_1d)), np.nan)
    for i, lat in enumerate(lats_1d):
        for j, lon in enumerate(lons_1d):
            pt = grid_lookup.get((lat, lon))
            if pt:
                val = value_func(pt)
                if val is not None:
                    Z[i, j] = val

    # Fill NaNs
    mask = np.isnan(Z)
    if mask.all():
        print(f"  Warning: no valid grid data for {mode} contour map")
        return
    if mask.any():
        Z[mask] = np.nanmean(Z)

    # Remove outlier spikes before interpolating (distance data has noisy routing)
    from scipy.ndimage import zoom, gaussian_filter, median_filter
    Z = median_filter(Z, size=3)  # kill spike outliers from routing quirks

    # Interpolate to a finer grid for smooth contours (40x40 → 240x240)
    interp_factor = 6
    Z = zoom(Z, interp_factor, order=1)  # linear interp (no ringing artifacts)
    sigma = 6 if mode == "distance" else 3  # distance needs more smoothing
    Z = gaussian_filter(Z, sigma=sigma)
    lats_1d = np.linspace(lats_1d[0], lats_1d[-1], Z.shape[0])
    lons_1d = np.linspace(lons_1d[0], lons_1d[-1], Z.shape[1])

    # Clamp values to ±10 range
    Z = np.clip(Z, -10, 10)

    levels = [-10, -5, -2, -1, 0, 1, 2, 5, 10]
    cmap = matplotlib.colormaps["RdBu"]
    norm = Normalize(vmin=-10, vmax=10)

    # ── Filled contour regions ──
    fig_tmp, ax_tmp = plt.subplots()
    cs_filled = ax_tmp.contourf(lons_1d, lats_1d, Z, levels=levels, cmap=cmap, norm=norm, extend="both")
    plt.close(fig_tmp)

    # Convert filled contour to GeoJSON polygons using allsegs
    for idx, segs in enumerate(cs_filled.allsegs):
        level_lo = levels[idx] if idx < len(levels) else levels[-1]
        level_hi = levels[idx + 1] if idx + 1 < len(levels) else levels[-1]
        level_mid = (level_lo + level_hi) / 2.0
        rgba = cmap(norm(level_mid))
        hex_color = "#{:02x}{:02x}{:02x}".format(int(rgba[0]*255), int(rgba[1]*255), int(rgba[2]*255))

        for seg in segs:
            coords = [[float(v[0]), float(v[1])] for v in seg]
            if len(coords) >= 3:
                if coords[0] != coords[-1]:
                    coords.append(coords[0])
                geojson = {
                    "type": "Feature",
                    "geometry": {"type": "Polygon", "coordinates": [coords]},
                    "properties": {"level_lo": level_lo, "level_hi": level_hi}
                }
                folium.GeoJson(
                    geojson,
                    style_function=lambda feat, c=hex_color: {
                        "fillColor": c, "color": c,
                        "weight": 0, "fillOpacity": 0.35
                    }
                ).add_to(m)

    # ── Contour lines ──
    fig_tmp, ax_tmp = plt.subplots()
    cs_lines = ax_tmp.contour(lons_1d, lats_1d, Z, levels=levels, colors="gray", linewidths=0.5)
    plt.close(fig_tmp)

    for idx, segs in enumerate(cs_lines.allsegs):
        level_val = levels[idx] if idx < len(levels) else 0
        for seg in segs:
            segment = [[float(v[1]), float(v[0])] for v in seg]  # [lat, lon]
            if len(segment) >= 2:
                line_color = "#333" if level_val == 0 else "#888"
                line_weight = 2.5 if level_val == 0 else 1.5
                folium.PolyLine(
                    segment,
                    color=line_color,
                    weight=line_weight,
                    dash_array="6 4" if level_val != 0 else None,
                    opacity=0.7,
                    tooltip=f"{level_val:+.0f} {unit_label} ({'Jackson closer' if level_val < 0 else 'Dalton closer' if level_val > 0 else 'Equal'})"
                ).add_to(m)

                # Label at midpoint of line
                if abs(level_val) in (0, 2, 5, 10):
                    mid = segment[len(segment) // 2]
                    label_text = f"{level_val:+.0f} {unit_label}" if level_val != 0 else f"Equal {title.split()[1]}"
                    folium.Marker(
                        location=mid,
                        icon=folium.DivIcon(html=f'<div style="font-size:10px;font-weight:bold;color:#333;'
                                                 f'background:rgba(255,255,255,0.8);padding:1px 4px;'
                                                 f'border-radius:2px;white-space:nowrap;">'
                                                 f'{label_text}</div>')
                    ).add_to(m)

    # ── Legend ──
    legend_html = f"""
    <div style="position:fixed;bottom:30px;left:30px;z-index:1000;
                background:white;padding:12px 16px;border-radius:8px;
                box-shadow:0 2px 6px rgba(0,0,0,0.3);font-size:13px;max-width:220px;">
    <b>{title}</b><br>
    <div style="display:flex;align-items:center;margin:6px 0;">
      <div style="width:120px;height:16px;background:linear-gradient(to right, #b2182b, #f4a582, #ffffff, #92c5de, #2166ac);border:1px solid #ccc;border-radius:2px;"></div>
    </div>
    <div style="display:flex;justify-content:space-between;font-size:10px;width:120px;">
      <span>Jackson</span><span>Equal</span><span>Dalton</span>
    </div>
    <div style="font-size:11px;margin-top:6px;">
    Contours at 0, ±1, ±2, ±5, ±10 {unit_label}<br>
    <span style="color:#333;">━━</span> Equal line<br>
    <span style="color:#888;">╌╌</span> Contour lines
    </div>
    </div>
    """
    m.get_root().html.add_child(folium.Element(legend_html))

    # Add markers last so they're on top of contour layers (hoverable)
    _add_church_markers(m)
    _add_address_markers(m, geocoded)

    folium.LayerControl().add_to(m)
    m.save(str(output_path))
    print(f"  {mode.capitalize()} contour map saved to {output_path}")


# ── Step 6: Generate line chart ────────────────────────────────────

def _generate_plotly_chart(geocoded, mode, output_path):
    """Generate scatter: X = to Jackson, Y = to Dalton. Diagonal = equal line."""

    if mode == "time":
        x_key, y_key = "time_jackson_min", "time_dalton_min"
        x_label = "Driving Time to Jackson (min)"
        y_label = "Driving Time to Dalton (min)"
        unit = "min"
        title_suffix = "Driving Time Comparison"
    else:
        x_key, y_key = "dist_jackson_mi", "dist_dalton_mi"
        x_label = "Driving Distance to Jackson (mi)"
        y_label = "Driving Distance to Dalton (mi)"
        unit = "mi"
        title_suffix = "Driving Distance Comparison"

    x_vals = [e[x_key] for e in geocoded]
    y_vals = [e[y_key] for e in geocoded]
    # Color: above diagonal (y > x) = Jackson closer (red), below = Dalton closer (blue)
    colors = ["#cc3333" if yv > xv else "#3333cc" if xv > yv else "#888"
              for xv, yv in zip(x_vals, y_vals)]

    hover_texts = []
    for e in geocoded:
        fam = ", ".join(e.get("families", ["Unknown"]))
        hover_texts.append(
            f"<b>{fam}</b><br>"
            f"{e['street']}, {e['city']}<br>"
            f"To Jackson: {e['dist_jackson_mi']} mi ({e['time_jackson_min']} min)<br>"
            f"To Dalton: {e['dist_dalton_mi']} mi ({e['time_dalton_min']} min)<br>"
            f"Closer to: {'Jackson' if e[y_key] > e[x_key] else 'Dalton'} "
            f"(by {abs(e[y_key] - e[x_key]):.1f} {unit})"
        )

    fig = go.Figure()

    # Scatter dots
    fig.add_trace(go.Scatter(
        x=x_vals, y=y_vals,
        mode="markers",
        marker=dict(size=7, color=colors, opacity=0.6,
                    line=dict(width=0.5, color="white")),
        text=hover_texts,
        hoverinfo="text",
        name="Family addresses",
    ))

    # Diagonal equal line
    lo = min(min(x_vals), min(y_vals)) - 2
    hi = max(max(x_vals), max(y_vals)) + 2
    fig.add_trace(go.Scatter(
        x=[lo, hi], y=[lo, hi],
        mode="lines",
        line=dict(color="#999", width=2, dash="dash"),
        name="Equal line",
        hoverinfo="skip",
    ))

    # Shade regions
    fig.add_trace(go.Scatter(
        x=[lo, hi, hi, lo], y=[lo, hi, hi + 10, lo + 10],
        fill="toself", fillcolor="rgba(204,51,51,0.04)",
        line=dict(width=0), mode="lines",
        name="Jackson closer (above)", hoverinfo="skip", showlegend=False,
    ))
    fig.add_trace(go.Scatter(
        x=[lo, hi, hi, lo], y=[lo, hi, hi - 10, lo - 10],
        fill="toself", fillcolor="rgba(51,51,204,0.04)",
        line=dict(width=0), mode="lines",
        name="Dalton closer (below)", hoverinfo="skip", showlegend=False,
    ))

    # Region labels
    fig.add_annotation(x=hi * 0.75, y=hi * 0.75 + (hi - lo) * 0.08,
                       text="Jackson closer", showarrow=False,
                       font=dict(size=12, color="#cc3333"), opacity=0.6)
    fig.add_annotation(x=hi * 0.75, y=hi * 0.75 - (hi - lo) * 0.08,
                       text="Dalton closer", showarrow=False,
                       font=dict(size=12, color="#3333cc"), opacity=0.6)

    # Annotate notable towns
    towns = {}
    for e in geocoded:
        city = e["city"]
        if city not in towns:
            towns[city] = {"x": [], "y": []}
        towns[city]["x"].append(e[x_key])
        towns[city]["y"].append(e[y_key])

    annotated = []
    for city, data in sorted(towns.items(), key=lambda x: -len(x[1]["x"])):
        if len(data["x"]) >= 3 and len(annotated) < 8:
            avg_x = float(np.mean(data["x"]))
            avg_y = float(np.mean(data["y"]))
            too_close = any(abs(avg_x - a[0]) < 2 and abs(avg_y - a[1]) < 2
                           for a in annotated)
            if not too_close:
                fig.add_annotation(
                    x=avg_x, y=avg_y, text=city,
                    showarrow=True, arrowhead=2, arrowsize=0.8,
                    arrowcolor="#aaa", font=dict(size=11, color="#444"),
                    ax=25, ay=-20,
                )
                annotated.append((avg_x, avg_y))

    jackson_n = sum(1 for xv, yv in zip(x_vals, y_vals) if yv > xv)
    dalton_n = sum(1 for xv, yv in zip(x_vals, y_vals) if xv > yv)

    fig.update_layout(
        title=dict(
            text=f"FBCJ Church Plant Analysis: {title_suffix}",
            subtitle=dict(
                text=(f"Each dot = one family. Above the diagonal = closer to Jackson, "
                      f"below = closer to Dalton. "
                      f"<b>{jackson_n}</b> closer to Jackson, "
                      f"<b>{dalton_n}</b> closer to Dalton by {unit}."),
                font=dict(size=12, color="#777"),
            ),
        ),
        xaxis_title=x_label,
        yaxis_title=y_label,
        xaxis=dict(scaleanchor="y", scaleratio=1),
        template="plotly_white",
        hovermode="closest",
        width=700, height=700,
        legend=dict(yanchor="top", y=0.99, xanchor="left", x=0.01),
    )

    fig.write_html(str(output_path), include_plotlyjs=True)
    print(f"  Chart saved to {output_path}")


def generate_charts(geocoded):
    """Generate both distance-based and time-based interactive Plotly charts."""
    print("\nGenerating interactive probability charts...")

    # Distance-based chart
    _generate_plotly_chart(geocoded, mode="distance", output_path=CHART_OUTPUT_DIST)

    # Time-based chart
    _generate_plotly_chart(geocoded, mode="time", output_path=CHART_OUTPUT_TIME)

    # Histogram of time differences
    _generate_histogram(geocoded)


def _generate_histogram(geocoded):
    """Generate a histogram with KDE curve overlay (no normality assumption)."""
    from scipy.stats import gaussian_kde, skew

    time_diffs = [e["time_diff"] for e in geocoded]
    n = len(time_diffs)
    mu = float(np.mean(time_diffs))
    median = float(np.median(time_diffs))
    sigma = float(np.std(time_diffs))
    skewness = float(skew(time_diffs))

    # Bin settings
    bin_size = 2
    min_val = int(min(time_diffs) // bin_size) * bin_size - bin_size
    max_val = int(max(time_diffs) // bin_size) * bin_size + bin_size * 2

    jackson_diffs = [d for d in time_diffs if d < 0]
    dalton_diffs = [d for d in time_diffs if d >= 0]

    within_5 = sum(1 for d in time_diffs if abs(d) <= 5)

    fig = go.Figure()

    # Red bars: Jackson closer
    fig.add_trace(go.Histogram(
        x=jackson_diffs,
        xbins=dict(start=min_val, end=0, size=bin_size),
        name=f"Jackson closer ({len(jackson_diffs)})",
        marker_color="#cc3333",
        opacity=0.75,
    ))

    # Blue bars: Dalton closer
    fig.add_trace(go.Histogram(
        x=dalton_diffs,
        xbins=dict(start=0, end=max_val, size=bin_size),
        name=f"Dalton closer ({len(dalton_diffs)})",
        marker_color="#3333cc",
        opacity=0.75,
    ))

    # KDE curve — follows the actual shape, no symmetry assumption
    kde = gaussian_kde(time_diffs)
    x_curve = np.linspace(min_val - 2, max_val + 2, 300)
    y_curve = kde(x_curve) * n * bin_size  # scale to match histogram counts
    fig.add_trace(go.Scatter(
        x=x_curve.tolist(), y=y_curve.tolist(),
        mode="lines",
        line=dict(color="#333", width=2.5),
        name="Density estimate (KDE)",
        hoverinfo="skip",
    ))

    # Equal driving time line
    fig.add_vline(x=0, line_dash="dash", line_color="#666", line_width=1.5,
                  annotation_text="Equal time", annotation_position="top right")

    # Median line (more robust than mean for skewed data)
    fig.add_vline(x=median, line_dash="dot", line_color="#333", line_width=1.5,
                  annotation_text=f"Median: {median:.1f} min",
                  annotation_position="top left")

    # Toss-up zone shading
    fig.add_vrect(x0=-5, x1=5, fillcolor="#ffee00", opacity=0.08,
                  layer="below", line_width=0)

    pct_jackson = len(jackson_diffs) / n * 100
    pct_dalton = len(dalton_diffs) / n * 100
    skew_dir = "toward Jackson" if skewness < 0 else "toward Dalton"

    # Count low-quality entries for note
    hq_n = sum(1 for e in geocoded if e.get("is_high_quality", True))
    lq_n = n - hq_n
    quality_note = f" ({lq_n} addresses use approx. data — see Excel for details)" if lq_n > 0 else ""

    fig.update_layout(
        title=dict(
            text="Distribution of Driving Time Differences",
            subtitle=dict(
                text=(f"N={n} families. Median = <b>{median:.1f} min</b>, "
                      f"Skewness = {skewness:.2f} ({skew_dir}). "
                      f"<b>{pct_jackson:.0f}%</b> closer to Jackson, "
                      f"<b>{pct_dalton:.0f}%</b> closer to Dalton. "
                      f"{within_5} families ({within_5/n*100:.0f}%) in the toss-up zone (within 5 min)."
                      f"{quality_note}"),
                font=dict(size=12, color="#777"),
            ),
        ),
        xaxis_title="Driving Time Difference: Jackson − Dalton (minutes)",
        yaxis_title="Number of Families",
        barmode="stack",
        template="plotly_white",
        width=1000, height=550,
        legend=dict(yanchor="top", y=0.99, xanchor="right", x=0.99),
        xaxis=dict(dtick=5),
    )

    fig.write_html(str(CHART_OUTPUT_HIST), include_plotlyjs=True)
    print(f"  Histogram saved to {CHART_OUTPUT_HIST}")


# ── Step 6b: Data quality and statistical analysis ────────────────

def flag_data_quality(geocoded):
    """Add data quality flags, household size, and adult/child counts to each entry."""
    for entry in geocoded:
        entry["is_approximate"] = entry.get("approximate", False)
        entry["is_fallback"] = abs(entry.get("time_jackson_min", 0) - entry.get("dist_jackson_mi", 0) * 2) < 0.2
        entry["is_high_quality"] = not entry["is_approximate"] and not entry["is_fallback"]

        # Use member details if available (from read_addresses), else fall back to family name count
        members = entry.get("members", [])
        if members:
            entry["adults"] = sum(1 for m in members if m["type"] == "adult")
            entry["graduated"] = sum(1 for m in members if m["type"] == "graduated")
            entry["children"] = sum(1 for m in members if m["type"] == "child")
            entry["household_size"] = len(members)
        else:
            member_count = sum(
                len([n for n in name_str.split(",") if n.strip()])
                for name_str in entry.get("families", ["?"])
            )
            entry["household_size"] = max(member_count, 1)
            entry["adults"] = entry["household_size"]
            entry["graduated"] = 0
            entry["children"] = 0

    hq = sum(1 for e in geocoded if e["is_high_quality"])
    lq = len(geocoded) - hq
    total_people = sum(e["household_size"] for e in geocoded)
    total_adults = sum(e["adults"] for e in geocoded)
    total_graduated = sum(e["graduated"] for e in geocoded)
    total_children = sum(e["children"] for e in geocoded)
    print(f"\nData quality: {hq} high-quality, {lq} low-quality (approx/fallback)")
    print(f"Total individuals: {total_people} ({total_adults} adults, {total_graduated} graduated, {total_children} children K-12)")
    return geocoded


def run_sensitivity_analysis(geocoded):
    """Test how conclusions change across different K values and incumbency bias levels."""
    k_values = [0.05, 0.08, 0.10, 0.15, 0.20, 0.25, 0.30]
    n = len(geocoded)

    print("\n" + "=" * 60)
    print("SENSITIVITY ANALYSIS")
    print("=" * 60)

    # ── Part 1: K parameter sensitivity ──
    print("\n  Part 1: K (logistic steepness) — controls how sharply time")
    print("  difference translates to probability.\n")
    print(f"  {'K':>6} | {'Lean Jackson':>14} | {'Toss-up':>14} | {'Lean Dalton':>14}")
    print("  " + "-" * 58)

    k_rows = []
    for k in k_values:
        j = d = t = 0
        for entry in geocoded:
            td = entry["time_jackson_min"] - entry["time_dalton_min"]
            p = 1.0 / (1.0 + math.exp(k * td))
            if p > 0.6:
                j += 1
            elif p < 0.4:
                d += 1
            else:
                t += 1
        star = " ◄" if k == K_LOGISTIC_TIME else ""
        print(f"  {k:.2f}   | {j:>4} ({j/n*100:5.1f}%) | {t:>4} ({t/n*100:5.1f}%) | {d:>4} ({d/n*100:5.1f}%){star}")
        k_rows.append({"k": k, "j_pct": j/n*100, "t_pct": t/n*100, "d_pct": d/n*100})

    # Robust findings
    always_j = sum(1 for e in geocoded if (e["time_jackson_min"] - e["time_dalton_min"]) < -15)
    always_d = sum(1 for e in geocoded if (e["time_jackson_min"] - e["time_dalton_min"]) > 15)
    swing = n - always_j - always_d
    print(f"\n  K-INDEPENDENT findings (hold at any K):")
    print(f"    {always_j} families ({always_j/n*100:.1f}%) are >15 min closer to Jackson")
    print(f"    {always_d} families ({always_d/n*100:.1f}%) are >15 min closer to Dalton")
    print(f"    {swing} families ({swing/n*100:.1f}%) are in the swing zone (K-dependent)")

    # ── Part 2: Incumbency bias sensitivity ──
    bias_values = [0, 5, 10, 15]
    print(f"\n  Part 2: Incumbency bias — how many extra minutes closer must Dalton")
    print(f"  be before it's a coin flip? (at K={K_LOGISTIC_TIME})\n")
    print(f"  {'Bias':>6} | {'Lean Jackson':>14} | {'Toss-up':>14} | {'Lean Dalton':>14} | Meaning")
    print("  " + "-" * 80)

    bias_rows = []
    for bias in bias_values:
        j = d = t = 0
        for entry in geocoded:
            td = entry["time_jackson_min"] - entry["time_dalton_min"]
            p = 1.0 / (1.0 + math.exp(K_LOGISTIC_TIME * (td - bias)))
            if p > 0.6:
                j += 1
            elif p < 0.4:
                d += 1
            else:
                t += 1
        meaning = "No loyalty factor" if bias == 0 else f"Need {bias} min advantage to switch"
        star = " ◄" if bias == 0 else ""
        print(f"  {bias:>4} min | {j:>4} ({j/n*100:5.1f}%) | {t:>4} ({t/n*100:5.1f}%) | {d:>4} ({d/n*100:5.1f}%) | {meaning}{star}")
        bias_rows.append({"bias": bias, "j_pct": j/n*100, "t_pct": t/n*100, "d_pct": d/n*100})

    # ── Generate Plotly sensitivity chart ──
    from plotly.subplots import make_subplots

    fig = make_subplots(rows=1, cols=2, subplot_titles=(
        "K Parameter Sensitivity", "Incumbency Bias Sensitivity"))

    # K sensitivity (left panel)
    for name, key, color in [("Lean Jackson (>60%)", "j_pct", "#cc3333"),
                              ("Toss-up (40-60%)", "t_pct", "#888888"),
                              ("Lean Dalton (<40%)", "d_pct", "#3333cc")]:
        fig.add_trace(go.Scatter(
            x=[r["k"] for r in k_rows], y=[r[key] for r in k_rows],
            mode="lines+markers", name=name,
            line=dict(color=color, width=2.5), marker=dict(size=8),
            legendgroup=name, showlegend=True,
        ), row=1, col=1)

    fig.add_vline(x=K_LOGISTIC_TIME, line_dash="dash", line_color="#333", line_width=1.5,
                  annotation_text=f"Current K={K_LOGISTIC_TIME}", row=1, col=1)

    # Bias sensitivity (right panel)
    for name, key, color in [("Lean Jackson (>60%)", "j_pct", "#cc3333"),
                              ("Toss-up (40-60%)", "t_pct", "#888888"),
                              ("Lean Dalton (<40%)", "d_pct", "#3333cc")]:
        fig.add_trace(go.Scatter(
            x=[r["bias"] for r in bias_rows], y=[r[key] for r in bias_rows],
            mode="lines+markers", name=name,
            line=dict(color=color, width=2.5), marker=dict(size=8),
            legendgroup=name, showlegend=False,
        ), row=1, col=2)

    fig.add_vline(x=0, line_dash="dash", line_color="#333", line_width=1.5,
                  annotation_text="Current (no bias)", row=1, col=2)

    fig.update_xaxes(title_text="K (logistic steepness)", row=1, col=1)
    fig.update_xaxes(title_text="Incumbency Bias (min)", row=1, col=2)
    fig.update_yaxes(title_text="% of Families", range=[0, 100], row=1, col=1)
    fig.update_yaxes(title_text="% of Families", range=[0, 100], row=1, col=2)

    fig.update_layout(
        title=dict(
            text="Sensitivity Analysis: How Model Parameters Affect Conclusions",
            subtitle=dict(
                text=(f"Left: varying K (steepness). Right: adding incumbency bias (switching cost). "
                      f"N={n} families. Robust findings: {always_j} always Jackson, "
                      f"{always_d} always Dalton regardless of parameters."),
                font=dict(size=12, color="#777"))),
        template="plotly_white",
        width=1100, height=500,
        legend=dict(yanchor="top", y=-0.15, xanchor="center", x=0.5, orientation="h"))

    output = SCRIPT_DIR / "FBCJ_Sensitivity_Analysis.html"
    fig.write_html(str(output), include_plotlyjs=True)
    print(f"\n  Sensitivity chart saved to {output.name}")


def compute_statistical_summary(geocoded):
    """Compute statistical tests on the time difference distribution."""
    from scipy.stats import wilcoxon

    try:
        from scipy.stats import binomtest
        def _binom_p(k, n):
            return binomtest(k, n, 0.5).pvalue
    except ImportError:
        from scipy.stats import binom_test
        def _binom_p(k, n):
            return binom_test(k, n, 0.5)

    hq = [e for e in geocoded if e.get("is_high_quality", True)]

    print("\n" + "=" * 60)
    print("STATISTICAL SUMMARY")
    print("=" * 60)

    for label, entries in [("All addresses", geocoded), ("High-quality only", hq)]:
        diffs = [e["time_diff"] for e in entries]
        n = len(diffs)
        if n < 5:
            continue

        median_val = float(np.median(diffs))
        mean_val = float(np.mean(diffs))
        std_val = float(np.std(diffs))

        # Wilcoxon signed-rank test: is the median time diff ≠ 0?
        stat, p_wilcox = wilcoxon(diffs)

        # Binomial test: is the closer-to-Jackson proportion ≠ 50%?
        n_jackson_closer = sum(1 for d in diffs if d < 0)
        p_binom = _binom_p(n_jackson_closer, n)

        # Effect size: matched-pairs rank-biserial correlation (r = Z / sqrt(N))
        # Standard effect size for the Wilcoxon signed-rank test (Rosenthal, 1991)
        z_score = (stat - n * (n + 1) / 4) / math.sqrt(n * (n + 1) * (2 * n + 1) / 24)
        effect_size = abs(z_score) / math.sqrt(n)
        effect_label = "small" if effect_size < 0.1 else "medium" if effect_size < 0.3 else "large"

        # Bootstrap 95% confidence intervals (10,000 resamples)
        rng = np.random.default_rng(seed=42)
        diffs_arr = np.array(diffs)
        n_boot = 10_000
        boot_medians = np.empty(n_boot)
        boot_jackson_pct = np.empty(n_boot)
        for b in range(n_boot):
            sample = rng.choice(diffs_arr, size=n, replace=True)
            boot_medians[b] = np.median(sample)
            boot_jackson_pct[b] = np.sum(sample < 0) / n * 100
        ci_median_lo, ci_median_hi = np.percentile(boot_medians, [2.5, 97.5])
        ci_jackson_lo, ci_jackson_hi = np.percentile(boot_jackson_pct, [2.5, 97.5])

        # Headcount-weighted stats
        total_people = sum(e.get("household_size", 1) for e in entries)
        people_jackson = sum(e.get("household_size", 1) for e in entries if e["time_diff"] < 0)
        people_dalton = total_people - people_jackson

        def _sig(p):
            return "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else "ns"

        direction = "Jackson closer" if median_val < 0 else "Dalton closer"
        jackson_pct = n_jackson_closer / n * 100

        print(f"\n  {label} (N={n}, {total_people} people):")
        print(f"    Median time diff:  {median_val:+.1f} min ({direction})  [95% CI: {ci_median_lo:+.1f} to {ci_median_hi:+.1f}]")
        print(f"    Mean time diff:    {mean_val:+.1f} min (SD={std_val:.1f})")
        print(f"    % closer Jackson:  {jackson_pct:.1f}%  [95% CI: {ci_jackson_lo:.1f}% to {ci_jackson_hi:.1f}%]")
        print(f"    Effect size (r):   {effect_size:.2f} ({effect_label})")
        print(f"    Wilcoxon test:     p = {p_wilcox:.4f} {_sig(p_wilcox)}")
        if p_wilcox < 0.05:
            print(f"      → The median time difference IS statistically significant.")
        else:
            print(f"      → The median time difference is NOT significant (could be chance).")
        print(f"    Binomial test:     {n_jackson_closer}/{n} closer to Jackson, p = {p_binom:.4f} {_sig(p_binom)}")
        if p_binom < 0.05:
            print(f"      → The split IS significantly different from 50/50.")
        else:
            print(f"      → The split is NOT significantly different from 50/50.")
        print(f"    By headcount:      {people_jackson} people ({people_jackson/total_people*100:.1f}%) closer to Jackson, "
              f"{people_dalton} ({people_dalton/total_people*100:.1f}%) closer to Dalton")

        # Adult/child breakdown
        adults_jackson = sum(e.get("adults", 0) for e in entries if e["time_diff"] < 0)
        adults_dalton = sum(e.get("adults", 0) for e in entries if e["time_diff"] >= 0)
        grad_jackson = sum(e.get("graduated", 0) for e in entries if e["time_diff"] < 0)
        grad_dalton = sum(e.get("graduated", 0) for e in entries if e["time_diff"] >= 0)
        kids_jackson = sum(e.get("children", 0) for e in entries if e["time_diff"] < 0)
        kids_dalton = sum(e.get("children", 0) for e in entries if e["time_diff"] >= 0)
        total_adults = adults_jackson + adults_dalton
        total_grad = grad_jackson + grad_dalton
        total_kids = kids_jackson + kids_dalton
        if total_adults > 0:
            print(f"    Adults:            {adults_jackson} ({adults_jackson/total_adults*100:.0f}%) Jackson, "
                  f"{adults_dalton} ({adults_dalton/total_adults*100:.0f}%) Dalton  (of {total_adults})")
        if total_grad > 0:
            print(f"    Graduated:         {grad_jackson} ({grad_jackson/total_grad*100:.0f}%) Jackson, "
                  f"{grad_dalton} ({grad_dalton/total_grad*100:.0f}%) Dalton  (of {total_grad})")
        if total_kids > 0:
            print(f"    Children (K-12):   {kids_jackson} ({kids_jackson/total_kids*100:.0f}%) Jackson, "
                  f"{kids_dalton} ({kids_dalton/total_kids*100:.0f}%) Dalton  (of {total_kids})")


def export_analysis_excel(geocoded):
    """Export full analysis results to Excel with formatting and summary."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    output = SCRIPT_DIR / "FBCJ_Analysis_Results.xlsx"
    wb = openpyxl.Workbook()

    # ── Sheet 1: Full data sorted by P(Dalton) descending ──
    ws = wb.active
    ws.title = "Analysis Results"

    headers = [
        "Family/Names", "Street", "City", "State", "Zip",
        "Household Size", "Adults", "Graduated", "Children (K-12)",
        "Dist to Jackson (mi)", "Time to Jackson (min)",
        "Dist to Dalton (mi)", "Time to Dalton (min)",
        "Distance Diff (mi)", "Time Diff (min)",
        "P(Jackson)", "P(Dalton)",
        "Most Likely Church", "Confidence Level",
        "Data Quality"
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    jackson_fill = PatternFill(start_color="FFDDE1", end_color="FFDDE1", fill_type="solid")
    dalton_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    tossup_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    low_quality_font = Font(italic=True, color="999999")
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"))

    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Sort: Dalton-leaning first, then toss-up, then Jackson
    sorted_data = sorted(geocoded, key=lambda e: e.get("p_dalton", 0), reverse=True)

    for row_idx, entry in enumerate(sorted_data, 2):
        families_str = ", ".join(entry.get("families", ["Unknown"]))
        p_j = entry.get("p_jackson", 0.5)
        p_d = entry.get("p_dalton", 0.5)

        if p_j > 0.8:
            likely, confidence = "Jackson", "Strong"
        elif p_j > 0.6:
            likely, confidence = "Jackson", "Moderate"
        elif p_d > 0.8:
            likely, confidence = "Dalton", "Strong"
        elif p_d > 0.6:
            likely, confidence = "Dalton", "Moderate"
        else:
            likely, confidence = "Toss-up", "Low"

        quality_issues = []
        if entry.get("is_approximate"):
            quality_issues.append("ZIP centroid")
        if entry.get("is_fallback"):
            quality_issues.append("Straight-line est.")
        quality = "High" if not quality_issues else "Low (" + ", ".join(quality_issues) + ")"

        row_data = [
            families_str,
            entry.get("street", ""),
            entry.get("city", ""),
            entry.get("state", ""),
            entry.get("zip", ""),
            entry.get("household_size", 1),
            entry.get("adults", 0),
            entry.get("graduated", 0),
            entry.get("children", 0),
            entry.get("dist_jackson_mi", ""),
            entry.get("time_jackson_min", ""),
            entry.get("dist_dalton_mi", ""),
            entry.get("time_dalton_min", ""),
            entry.get("dist_diff", ""),
            entry.get("time_diff", ""),
            p_j,
            p_d,
            likely,
            confidence,
            quality,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

            if likely == "Jackson":
                cell.fill = jackson_fill
            elif likely == "Dalton":
                cell.fill = dalton_fill
            elif likely == "Toss-up":
                cell.fill = tossup_fill

            if quality.startswith("Low"):
                cell.font = low_quality_font

        # Format probability columns as percentage (P(Jackson)=col 16, P(Dalton)=col 17)
        ws.cell(row=row_idx, column=16).number_format = '0%'
        ws.cell(row=row_idx, column=17).number_format = '0%'

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, min(len(sorted_data) + 2, 50)):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Summary ──
    ws2 = wb.create_sheet("Summary")

    jackson_entries = [e for e in geocoded if e.get("p_jackson", 0.5) > 0.6]
    dalton_entries = [e for e in geocoded if e.get("p_dalton", 0.5) > 0.6]
    tossup_entries = [e for e in geocoded if 0.4 <= e.get("p_jackson", 0.5) <= 0.6]
    n = len(geocoded)

    hq = [e for e in geocoded if e.get("is_high_quality", True)]
    hq_j = sum(1 for e in hq if e.get("p_jackson", 0.5) > 0.6)
    hq_d = sum(1 for e in hq if e.get("p_dalton", 0.5) > 0.6)
    hq_t = len(hq) - hq_j - hq_d

    total_people = sum(e.get("household_size", 1) for e in geocoded)
    people_j = sum(e.get("household_size", 1) for e in jackson_entries)
    people_d = sum(e.get("household_size", 1) for e in dalton_entries)
    people_t = total_people - people_j - people_d

    # Adult/child/graduated breakdown per classification
    def _sum_type(entries, field):
        return sum(e.get(field, 0) for e in entries)

    total_adults = _sum_type(geocoded, "adults")
    total_grad = _sum_type(geocoded, "graduated")
    total_kids = _sum_type(geocoded, "children")
    adults_j = _sum_type(jackson_entries, "adults")
    adults_d = _sum_type(dalton_entries, "adults")
    adults_t = total_adults - adults_j - adults_d
    grad_j = _sum_type(jackson_entries, "graduated")
    grad_d = _sum_type(dalton_entries, "graduated")
    grad_t = total_grad - grad_j - grad_d
    kids_j = _sum_type(jackson_entries, "children")
    kids_d = _sum_type(dalton_entries, "children")
    kids_t = total_kids - kids_j - kids_d

    def _pct(num, denom):
        return f"{num/denom*100:.1f}%" if denom > 0 else "N/A"

    summary_rows = [
        ("FBCJ Church Plant Analysis — Summary", "", True),
        ("", "", False),
        ("ADDRESSES ANALYZED", "", True),
        ("Total addresses", f"{n}"),
        ("High-quality data", f"{len(hq)}"),
        ("Low-quality (approx/fallback)", f"{n - len(hq)}"),
        ("", "", False),
        ("ALL ADDRESSES — BY ADDRESS", "", True),
        ("Lean Jackson (P>60%)", f"{len(jackson_entries)} ({len(jackson_entries)/n*100:.1f}%)"),
        ("Lean Dalton (P>60%)", f"{len(dalton_entries)} ({len(dalton_entries)/n*100:.1f}%)"),
        ("Toss-up (P 40-60%)", f"{len(tossup_entries)} ({len(tossup_entries)/n*100:.1f}%)"),
        ("", "", False),
        ("HIGH-QUALITY ONLY — BY ADDRESS", "", True),
        ("Lean Jackson", f"{hq_j} ({hq_j/len(hq)*100:.1f}%)" if hq else "N/A"),
        ("Lean Dalton", f"{hq_d} ({hq_d/len(hq)*100:.1f}%)" if hq else "N/A"),
        ("Toss-up", f"{hq_t} ({hq_t/len(hq)*100:.1f}%)" if hq else "N/A"),
        ("", "", False),
        ("BY HEADCOUNT — ALL PEOPLE", "", True),
        ("Total people", f"{total_people}"),
        ("People leaning Jackson", f"{people_j} ({_pct(people_j, total_people)})"),
        ("People leaning Dalton", f"{people_d} ({_pct(people_d, total_people)})"),
        ("People in toss-up", f"{people_t} ({_pct(people_t, total_people)})"),
        ("", "", False),
        ("BY HEADCOUNT — ADULTS ONLY", "", True),
        ("Total adults", f"{total_adults}"),
        ("Adults leaning Jackson", f"{adults_j} ({_pct(adults_j, total_adults)})"),
        ("Adults leaning Dalton", f"{adults_d} ({_pct(adults_d, total_adults)})"),
        ("Adults in toss-up", f"{adults_t} ({_pct(adults_t, total_adults)})"),
        ("", "", False),
        ("BY HEADCOUNT — GRADUATED (young adults)", "", True),
        ("Total graduated", f"{total_grad}"),
        ("Graduated leaning Jackson", f"{grad_j} ({_pct(grad_j, total_grad)})"),
        ("Graduated leaning Dalton", f"{grad_d} ({_pct(grad_d, total_grad)})"),
        ("Graduated in toss-up", f"{grad_t} ({_pct(grad_t, total_grad)})"),
        ("", "", False),
        ("BY HEADCOUNT — CHILDREN (K-12)", "", True),
        ("Total children", f"{total_kids}"),
        ("Children leaning Jackson", f"{kids_j} ({_pct(kids_j, total_kids)})"),
        ("Children leaning Dalton", f"{kids_d} ({_pct(kids_d, total_kids)})"),
        ("Children in toss-up", f"{kids_t} ({_pct(kids_t, total_kids)})"),
        ("", "", False),
        ("MODEL PARAMETERS", "", True),
        ("K (logistic steepness)", f"{K_LOGISTIC_TIME}"),
        ("Incumbency bias", "0 (not applied — see sensitivity analysis)"),
        ("Driving times from", "Google Maps Distance Matrix API"),
        ("Departure time", "Next Sunday, 9:00 AM EDT"),
        ("Outlier filter", "35 mi from church midpoint"),
        ("", "", False),
        ("METHODOLOGY NOTES", "", True),
        ("Model", "P(Jackson) = 1 / (1 + exp(K × (time_to_Jackson − time_to_Dalton)))"),
        ("Interpretation", f"At K={K_LOGISTIC_TIME}, a 10-min time difference → ~{1/(1+math.exp(K_LOGISTIC_TIME*10)):.0%} probability for the farther church"),
        ("Sensitivity", "See FBCJ_Sensitivity_Analysis.html for how K and incumbency bias affect results"),
        ("Timing note", "Driving times use EDT departure; EST months may differ by 1-2 min"),
        ("Low-quality data", "ZIP centroid entries use approximate location; straight-line entries used 1.3x road factor"),
        ("Limitation", "This model uses driving time ONLY — does not account for social ties, loyalty, or ministry involvement"),
        ("Limitation", "Incumbency bias of 0 means equidistant = 50/50, which overstates switching likelihood"),
    ]

    for row_idx, row_data in enumerate(summary_rows, 1):
        label = row_data[0]
        value = row_data[1] if len(row_data) > 1 else ""
        is_header = row_data[2] if len(row_data) > 2 else False

        ws2.cell(row=row_idx, column=1, value=label)
        ws2.cell(row=row_idx, column=2, value=value)
        if is_header:
            ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=12 if row_idx == 1 else 11)

    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['B'].width = 85

    wb.save(output)
    print(f"\n  Excel export saved to {output.name}")
    print(f"    Sheet 1: Full results (sorted by P(Dalton) desc — top = most likely Dalton)")
    print(f"    Sheet 2: Summary with methodology notes")


# ── Main ───────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("FBCJ Church Plant Analysis: Jackson vs Dalton")
    print("=" * 60)

    # Step 1: Read addresses
    addresses, members_by_address = read_addresses()

    # Step 2: Geocode
    geocoded = geocode_addresses(addresses, members_by_address)
    print(f"Successfully geocoded {len(geocoded)} addresses")

    # Step 3: Calculate driving distances
    geocoded = calculate_distances(geocoded)

    # Step 3b: Retry any addresses that fell back to straight-line estimates
    geocoded = retry_fallback_distances(geocoded)

    # Step 3c: Filter geographic outliers
    geocoded = filter_geographic_outliers(geocoded, max_miles=35)

    # Step 4: Calculate probabilities
    geocoded = calculate_probabilities(geocoded)

    # Step 4b: Flag data quality and compute household sizes
    geocoded = flag_data_quality(geocoded)

    # Step 4c: Statistical summary (with fallback exclusion)
    compute_statistical_summary(geocoded)

    # Step 4d: Sensitivity analysis
    run_sensitivity_analysis(geocoded)

    # Step 5: Generate map
    generate_map(geocoded)

    # Step 5b: Generate contour maps
    generate_contour_map(geocoded, "time")
    generate_contour_map(geocoded, "distance")

    # Step 6: Generate charts (distance + time)
    generate_charts(geocoded)

    # Step 7: Export analysis to Excel
    export_analysis_excel(geocoded)

    print("\n" + "=" * 60)
    print("Done! Open in your browser:")
    print(f"  - {MAP_OUTPUT.name} (combined boundary map)")
    print(f"  - {MAP_OUTPUT_TIME.name} (time contour map)")
    print(f"  - {MAP_OUTPUT_DIST.name} (distance contour map)")
    print(f"  - FBCJ_Sensitivity_Analysis.html (K & bias sensitivity)")
    print(f"  - FBCJ_Analysis_Results.xlsx (full Excel export)")
    print("=" * 60)


if __name__ == "__main__":
    main()
